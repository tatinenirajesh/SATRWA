"""SATRWA Iteration 3 tests: UPI info+QR endpoints, admin verify-payment,
receipt has verified/upi_ref_no, PDF stamp reflects verified status,
Excel export contains verified + upi_ref_no columns."""
import io
import os
import uuid
from datetime import datetime, timezone

import openpyxl
import pytest
import requests

BASE_URL = os.environ.get(
    "EXPO_PUBLIC_BACKEND_URL",
    "https://township-maintenance.preview.emergentagent.com",
).rstrip("/")
API = f"{BASE_URL}/api"
ADMIN_PIN = "1234"

RUN = uuid.uuid4().hex[:6]
FLAT = {"block": "A", "flat_no": f"U{RUN}"}


def current_ym():
    n = datetime.now(timezone.utc)
    return f"{n.year:04d}-{n.month:02d}"


@pytest.fixture(scope="module")
def client():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    return s


@pytest.fixture(scope="module", autouse=True)
def _register_flat(client):
    r = client.post(f"{API}/auth/register", json={
        **FLAT, "bhk_type": "2BHK", "owner_name": "TEST_UPI",
        "phone": "9000000000", "start_month": current_ym(),
    })
    assert r.status_code in (200, 400)  # may already exist


# ---------------- UPI /info ----------------
class TestUpiInfo:
    def test_upi_info_fields(self, client):
        r = client.get(f"{API}/upi/info", params={"amount": 2000, "note": "Maint A-101"})
        assert r.status_code == 200
        d = r.json()
        assert d["vpa"] == "satrwa@icici"
        assert d["payee_name"] == "Sri Anjaneya Township RWA"
        assert d["upi_url"].startswith("upi://pay?pa=satrwa@icici")
        # required params in upi_url
        for tok in ["pa=satrwa%40icici" ,"pa=satrwa@icici"]:
            pass
        for req in ["pa=", "pn=", "am=2000.00", "cu=INR", "tn="]:
            assert req in d["upi_url"], f"{req} missing in upi_url"
        # app-specific schemes
        assert d["gpay_url"].startswith("tez://")
        assert d["phonepe_url"].startswith("phonepe://")
        assert d["paytm_url"].startswith("paytmmp://")
        for u in (d["gpay_url"], d["phonepe_url"], d["paytm_url"]):
            for req in ["pa=", "pn=", "am=", "cu=INR", "tn="]:
                assert req in u


# ---------------- UPI /qr ----------------
class TestUpiQr:
    def test_qr_returns_png(self, client):
        r = client.get(f"{API}/upi/qr", params={"amount": 1500, "note": "Gym"})
        assert r.status_code == 200
        assert r.headers.get("content-type", "").startswith("image/png")
        assert r.content.startswith(b"\x89PNG\r\n\x1a\n")
        assert len(r.content) > 500  # sanity


# ------------- Payment stores upi_ref_no + verified=False -------------
class TestPaymentVerifiedFields:
    def test_maintenance_pay_stores_upi_ref_no(self, client):
        r = client.post(f"{API}/maintenance/pay", json={
            **FLAT, "mode": "full", "include_conveyance": False,
            "upi_id": "payer@upi", "upi_ref_no": "501234567890",
        })
        assert r.status_code == 200, r.text
        rec = r.json()["receipt"]
        assert rec["upi_ref_no"] == "501234567890"
        assert rec["verified"] is False
        # confirm persisted via GET
        g = client.get(f"{API}/receipt/{rec['receipt_no']}")
        assert g.status_code == 200
        gd = g.json()
        assert gd["upi_ref_no"] == "501234567890"
        assert gd["verified"] is False
        pytest.MAINT_RECEIPT = rec["receipt_no"]

    def test_gym_booking_stores_upi_ref_no(self, client):
        r = client.post(f"{API}/amenity/gym", json={
            **FLAT, "members": 2, "booking_date": "2026-07-15",
            "upi_id": "x@upi", "upi_ref_no": "600000000001",
        })
        assert r.status_code == 200, r.text
        rec = r.json()["receipt"]
        assert rec["upi_ref_no"] == "600000000001"
        assert rec["verified"] is False
        pytest.GYM_RECEIPT = rec["receipt_no"]

    def test_pool_booking_stores_upi_ref_no(self, client):
        r = client.post(f"{API}/amenity/pool", json={
            **FLAT, "persons": 2, "booking_date": "2026-07-15",
            "upi_id": "x@upi", "upi_ref_no": "700000000002",
        })
        assert r.status_code == 200, r.text
        rec = r.json()["receipt"]
        assert rec["upi_ref_no"] == "700000000002"
        assert rec["verified"] is False


# --------------- Admin verify-payment ---------------
class TestAdminVerifyPayment:
    def test_wrong_pin_returns_401(self, client):
        r = client.post(f"{API}/admin/verify-payment", json={
            "receipt_no": pytest.MAINT_RECEIPT, "pin": "0000", "verified": True,
        })
        assert r.status_code == 401

    def test_missing_receipt_returns_404(self, client):
        r = client.post(f"{API}/admin/verify-payment", json={
            "receipt_no": "NOPE999", "pin": ADMIN_PIN, "verified": True,
        })
        assert r.status_code == 404

    def test_verify_toggles_flag(self, client):
        # verify true
        r = client.post(f"{API}/admin/verify-payment", json={
            "receipt_no": pytest.MAINT_RECEIPT, "pin": ADMIN_PIN, "verified": True,
        })
        assert r.status_code == 200
        assert r.json()["verified"] is True
        # confirm via GET
        g = client.get(f"{API}/receipt/{pytest.MAINT_RECEIPT}").json()
        assert g["verified"] is True
        assert g.get("verified_at")

        # undo
        r2 = client.post(f"{API}/admin/verify-payment", json={
            "receipt_no": pytest.MAINT_RECEIPT, "pin": ADMIN_PIN, "verified": False,
        })
        assert r2.status_code == 200
        assert r2.json()["verified"] is False
        g2 = client.get(f"{API}/receipt/{pytest.MAINT_RECEIPT}").json()
        assert g2["verified"] is False

    def test_amenity_receipt_can_be_verified(self, client):
        r = client.post(f"{API}/admin/verify-payment", json={
            "receipt_no": pytest.GYM_RECEIPT, "pin": ADMIN_PIN, "verified": True,
        })
        assert r.status_code == 200
        g = client.get(f"{API}/receipt/{pytest.GYM_RECEIPT}").json()
        assert g["verified"] is True


# --------------- PDF reflects verified status ---------------
class TestPdfStamp:
    def test_pdf_pending_and_paid(self, client):
        # currently maint is unverified again
        r1 = client.get(f"{API}/receipt/{pytest.MAINT_RECEIPT}/pdf")
        assert r1.status_code == 200
        assert r1.content.startswith(b"%PDF")
        assert len(r1.content) > 500
        # verify then re-fetch
        client.post(f"{API}/admin/verify-payment", json={
            "receipt_no": pytest.MAINT_RECEIPT, "pin": ADMIN_PIN, "verified": True,
        })
        r2 = client.get(f"{API}/receipt/{pytest.MAINT_RECEIPT}/pdf")
        assert r2.status_code == 200
        assert r2.content.startswith(b"%PDF")
        # different content since stamp text changes PENDING->PAID
        assert r1.content != r2.content


# --------------- Excel has verified + upi_ref_no columns ---------------
class TestExcelColumns:
    def test_export_has_new_columns(self, client):
        r = client.get(f"{API}/admin/export", params={"pin": ADMIN_PIN})
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert "Maintenance" in wb.sheetnames
        assert "Amenities" in wb.sheetnames
        m_headers = [c.value for c in wb["Maintenance"][1]]
        a_headers = [c.value for c in wb["Amenities"][1]]
        assert "verified" in m_headers
        assert "upi_ref_no" in m_headers
        assert "verified" in a_headers
        assert "upi_ref_no" in a_headers
